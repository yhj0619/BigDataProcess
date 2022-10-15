#!/usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook( "student.xlsx" )
ws = wb['Sheet1']

row_id = 1
total_list = []
total_sort = []
grade = []

for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row=row_id,column = 3).value*0.3
		sum_v += ws.cell(row=row_id,column=4).value*0.35
		sum_v += ws.cell(row=row_id,column=5).value*0.34
		sum_v += ws.cell(row=row_id,column=6).value
		ws.cell(row=row_id,column=7).value = sum_v
		total_list.append(sum_v)
	row_id += 1

total_sort=sorted(total_list,reverse=True)


a_grade = 0
b_grade = 0
c_grade = 0
i = 0
stu_num = len(total_sort)*0.3
while total_sort:
	if stu_num-total_sort.count(total_sort[i]) >= 0:
		row_new = 2
		count = 0
		for row in ws:
			if ws.cell(row = row_new, column = 7).value == total_sort[i]:
				ws.cell(row=row_new, column = 8).value='A0'
				grade.append('A0')
				stu_num -= 1
				count += 1
				a_grade += 1
			row_new += 1
		i += count	
	else:
		break
stu_num = stu_num + len(total_sort) * 0.4
while total_sort:
	if stu_num - total_sort.count(total_sort[i]) >= 0:
		row_new = 2
		count = 0
		for row in ws:	
			if ws.cell(row=row_new, column=7).value == total_sort[i]:
				ws.cell(row=row_new,column=8).value='B0'	
				grade.append('B0')
				stu_num -= 1
				count += 1
				b_grade += 1
			row_new +=1
		i += count
	else:
		break
while i < len(total_sort):	
	row_new = 2
	count = 0
	for row in ws:
		if ws.cell(row=row_new, column=7).value == total_sort[i]:
			ws.cell(row=row_new, column=8).value = 'C0'
			grade.append('C0')
			count += 1
			c_grade += 1
		row_new += 1
	i += count
i = 0
betterA_num = a_grade * 0.5
while total_sort:
	if betterA_num-total_sort.count(total_sort[i]) >= 0:
		row_new = 2
		count = 0
		for row in ws:
			if ws.cell(row = row_new, column = 7).value == total_sort[i]:
				ws.cell(row=row_new, column = 8).value='A+'
				grade[i] = 'A+'
				betterA_num -= 1
				count += 1
			row_new += 1
		i += count	
	else:
		break
i=a_grade
betterB_num = b_grade * 0.5
while total_sort:
	if betterB_num-total_sort.count(total_sort[i]) >= 0:
		row_new = 2
		count = 0
		for row in ws:
			if ws.cell(row = row_new, column = 7).value == total_sort[i]:
				ws.cell(row=row_new, column = 8).value='B+'
				grade[i] = 'B+'
				betterB_num -= 1
				count += 1
			row_new += 1
		i += count	
	else:
		break

i = a_grade + b_grade
betterC_num = c_grade * 0.5
while total_sort:
	if betterC_num-total_sort.count(total_sort[i]) >= 0:
		row_new = 2
		count = 0
		for row in ws:
			if ws.cell(row = row_new, column = 7).value == total_sort[i]:
				ws.cell(row=row_new, column = 8).value='C+'
				grade[i] = 'C+'
				betterC_num -= 1
				count += 1
			row_new += 1
		i += count	
	else:
		break

wb.save( "student.xlsx" )
